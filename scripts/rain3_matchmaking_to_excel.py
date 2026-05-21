#!/usr/bin/env python3
"""
Delllo RAIN3.0 — CIL 100-User Matchmaking Test
═══════════════════════════════════════════════════════════════════
Reads cil_population_100_users.xlsx, creates a new tenant,
ingests all 100 users, runs matchmaking for each, and saves
all results to matchmaking_results.xlsx.

Steps:
  1  Wipe / create fresh CIL tenant
  2  Create 100 users from Excel dataset
  3  Ingest CV text for every user (built from CIL profile fields)
  4  Run LLM extraction for every user (iKG population)
  5  Post intent signals where primary_driver exists
  6  Run matchmaking for every user
  7  Save all matches + score breakdowns to Excel

Usage:
  pip install pandas openpyxl httpx
  python test_cil_matchmaking.py

Requirements:
  - docker compose up (postgres, memgraph, minio)
  - ollama serve + model pulled
  - uvicorn app.main:app running on port 8000
═══════════════════════════════════════════════════════════════════
"""

import sys
import json
import time
import uuid
import httpx
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    import psycopg2
    import psycopg2.extras
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False

# ─────────────────────────────────────────────
#  Config
# ─────────────────────────────────────────────

BASE_URL    = "http://localhost:8000"
TENANT_ID   = "11000000-0000-0000-0000-000000000001"
TENANT_NAME = "CIL Test Tenant"
TENANT_SLUG = "cil-test-100"

# Direct Postgres connection (bypasses Ollama for fact seeding)
PG_DSN = "postgresql://delllo:delllo_secret@localhost:5432/delllo_db"

SCRIPT_DIR  = Path(__file__).parent.resolve()
ROOT_DIR    = SCRIPT_DIR.parent
INPUT_FILE  = ROOT_DIR / "cil_population_100_users.xlsx"
OUTPUT_FILE = ROOT_DIR / "matchmaking_results.xlsx"

TX_TYPE     = "knowledge_transfer"
MAX_MATCHES = 10

TIMEOUT_INGEST = 60
TIMEOUT_MATCH  = 120

# ─────────────────────────────────────────────
#  Terminal helpers
# ─────────────────────────────────────────────

PASS = "✓"
FAIL = "✗"
WARN = "⚠"
results_log = []

def section(title):
    print(f"\n{'═'*64}")
    print(f"  {title}")
    print(f"{'═'*64}")

def info(msg):
    print(f"     {msg}")

def check(condition, msg, fatal=False):
    icon = PASS if condition else FAIL
    print(f"  {icon}  {msg}")
    results_log.append((condition, msg))
    if not condition and fatal:
        print("\n  ❌ Fatal — stopping.")
        sys.exit(1)

# ─────────────────────────────────────────────
#  Build CV text from CIL profile
# ─────────────────────────────────────────────

def build_cv_text(row):
    """Convert a CIL Excel row into a rich CV text for ingestion."""
    lines = []

    # Header
    lines.append(f"{row['full_name']} — {row['current_role']}, {row['organisation_name']}")
    lines.append("")

    # Identity
    lines.append("IDENTITY")
    lines.append(f"Organisation: {row['organisation_name']} ({row['organisation_archetype']})")
    lines.append(f"Location: {row['location']}")
    lines.append(f"Seniority: {row['seniority']}")
    lines.append(f"Role Family: {row['role_family']}")
    lines.append(f"Persona: {row['persona_cluster']}")
    lines.append(f"Market Regime: {row['market_regime']}")
    lines.append("")

    # Employment + Education
    try:
        hist = json.loads(row['previous_employment_and_education_history_json']) if pd.notna(row['previous_employment_and_education_history_json']) else {}
        emp = hist.get("employment_history", [])
        edu = hist.get("education_history", [])
        if emp:
            lines.append("EXPERIENCE")
            for e in emp:
                lines.append(f"- {e.get('role','?')} at {e.get('organisation','?')} ({e.get('period','?')})")
            lines.append("")
        if edu:
            lines.append("EDUCATION")
            for e in edu:
                lines.append(f"- {e.get('qualification','?')} in {e.get('field','?')} from {e.get('institution','?')}")
            lines.append("")
    except Exception:
        pass

    # Drivers (needs/objectives context)
    lines.append("MARKET FOCUS")
    lines.append(f"Primary Driver: {row['primary_driver_description']}")
    lines.append(f"Secondary Driver: {row['secondary_driver_description']}")
    lines.append("")

    # Offers
    try:
        offers = json.loads(row['offers_json']) if pd.notna(row['offers_json']) else []
        if offers:
            lines.append("OFFERS")
            for o in offers:
                lines.append(f"- [{o.get('offer_type','?')}] {o.get('summary','')}")
            lines.append("")
    except Exception:
        pass

    # Needs
    try:
        needs = json.loads(row['needs_json']) if pd.notna(row['needs_json']) else []
        if needs:
            lines.append("NEEDS")
            for n in needs:
                lines.append(f"- [{n.get('need_type','?')}] {n.get('summary','')}")
            lines.append("")
    except Exception:
        pass

    # Insights
    try:
        insights = json.loads(row['insights_json']) if pd.notna(row['insights_json']) else []
        if insights:
            lines.append("INSIGHTS AND MARKET VIEWS")
            for i in insights:
                lines.append(f"- [{i.get('insight_type','?')}] {i.get('summary','')}")
            lines.append("")
    except Exception:
        pass

    # Solutions
    try:
        solutions = json.loads(row['solutions_json']) if pd.notna(row['solutions_json']) else []
        if solutions:
            lines.append("SOLUTIONS AND EXPERTISE")
            for s in solutions:
                prereqs = ", ".join(s.get("prerequisites", []))
                lines.append(f"- {s.get('summary','')} (prerequisites: {prereqs})")
            lines.append("")
    except Exception:
        pass

    return "\n".join(lines)


# ─────────────────────────────────────────────
#  Step 0 — Wipe + create tenant
# ─────────────────────────────────────────────

def step_setup_tenant():
    section("Step 0 — Setup CIL Tenant")

    with httpx.Client(timeout=30) as c:
        # Health check
        r = c.get(f"{BASE_URL}/health")
        check(r.status_code == 200, "API is healthy", fatal=True)

        # Wipe existing tenant data (safe to fail if tenant doesn't exist yet)
        r = c.post(f"{BASE_URL}/v1/admin/wipe",
                   json={"tenant_id": TENANT_ID, "confirm": True})
        if r.status_code == 200:
            info("Existing tenant data wiped")
        else:
            info("No existing data to wipe (first run)")

        info(f"Tenant {TENANT_ID} ({TENANT_SLUG}) ready")


# ─────────────────────────────────────────────
#  Step 1 — Create 100 users
# ─────────────────────────────────────────────

def step_create_users(df):
    section("Step 1 — Create 100 Users")
    user_ids = {}

    with httpx.Client(timeout=30) as c:
        for idx, row in df.iterrows():
            uid = str(uuid.uuid5(uuid.UUID(TENANT_ID), row['full_name']))
            email = f"{row['full_name'].lower().replace(' ', '.').replace('/', '')}@cil-test.com"

            r = c.post(f"{BASE_URL}/v1/users", json={
                "user_id":      uid,
                "tenant_id":    TENANT_ID,
                "display_name": row['full_name'],
                "email":        email,
                "headline":     f"{row['current_role']} at {row['organisation_name']}",
                "role":         "member",
                "status":       "active",
            })
            ok = r.status_code in (200, 201)
            user_ids[row['full_name']] = uid
            if (idx + 1) % 10 == 0:
                check(ok, f"Users created: {idx + 1}/100")

    info(f"Total users created: {len(user_ids)}")
    return user_ids


# ─────────────────────────────────────────────
#  Step 2 — Ingest CV text for all users
# ─────────────────────────────────────────────

def step_ingest(df, user_ids):
    section("Step 2 — Ingest CV Profiles")
    doc_ids = {}
    failed = 0

    with httpx.Client(timeout=TIMEOUT_INGEST) as c:
        for idx, row in df.iterrows():
            name = row['full_name']
            uid  = user_ids[name]
            cv_text = build_cv_text(row)

            r = c.post(f"{BASE_URL}/v1/ingest/text", data={
                "tenant_id":    TENANT_ID,
                "user_id":      uid,
                "content":      cv_text,
                "source_type":  "cv",
                "filename":     f"{name.replace(' ','_')}_cil_profile.txt",
                "embed":        "true",
            })

            if r.status_code == 200:
                doc_ids[name] = r.json()["document_id"]
            else:
                failed += 1
                info(f"  WARN: Ingest failed for {name}: {r.status_code}")

            if (idx + 1) % 10 == 0:
                print(f"  →  Ingested {idx + 1}/100 profiles...")

    check(failed == 0, f"All ingestions succeeded (failed: {failed})")
    info(f"Documents created: {len(doc_ids)}")
    return doc_ids


# ─────────────────────────────────────────────
#  Step 3 — Extract facts (iKG population)
# ─────────────────────────────────────────────
#  Step 3 — Seed facts directly into Postgres
#  Bypasses Ollama entirely — uses structured
#  CIL JSON fields from the Excel as fact source
# ─────────────────────────────────────────────

def _parse_json_col(val):
    if pd.isna(val): return []
    try:    return json.loads(val)
    except: return []

def _slug(text):
    """Make a canonical snake_case key from free text."""
    import re
    return re.sub(r'[^a-z0-9]+', '_', str(text).lower()).strip('_')[:120]

def build_facts_from_row(row, user_id):
    """
    Parse all structured CIL JSON columns and return a list of
    (fact_type, canonical_value, raw_value, confidence, visibility) tuples.
    """
    facts = []
    tid   = TENANT_ID
    vis   = "match_engine_only"

    # ── Offers → 'offer' fact type ───────────────────────────
    for o in _parse_json_col(row.get('offers_json')):
        raw = o.get('summary', '')
        if raw:
            facts.append(('offer', _slug(raw), raw, 0.90, vis))

    # ── Needs → 'need' fact type ─────────────────────────────
    for n in _parse_json_col(row.get('needs_json')):
        raw = n.get('summary', '')
        if raw:
            facts.append(('need', _slug(raw), raw, 0.90, vis))

    # ── Solutions → 'skill' fact type ────────────────────────
    for s in _parse_json_col(row.get('solutions_json')):
        raw = s.get('summary', '')
        if raw:
            facts.append(('skill', _slug(raw), raw, 0.85, vis))

    # ── Insights → 'topic' fact type ─────────────────────────
    for i in _parse_json_col(row.get('insights_json')):
        raw = i.get('summary', '')
        if raw:
            facts.append(('topic', _slug(raw), raw, 0.80, vis))

    # ── Primary driver → 'domain' ────────────────────────────
    pd_desc = str(row.get('primary_driver_description', ''))
    if pd_desc:
        facts.append(('domain', _slug(row.get('primary_driver_id', pd_desc)),
                      pd_desc, 0.95, vis))

    # ── Secondary driver → 'domain' ──────────────────────────
    sd_desc = str(row.get('secondary_driver_description', ''))
    if sd_desc and sd_desc != pd_desc:
        facts.append(('domain', _slug(row.get('secondary_driver_id', sd_desc)),
                      sd_desc, 0.85, vis))

    # ── Persona cluster → 'topic' ────────────────────────────
    persona = str(row.get('persona_cluster', ''))
    if persona:
        facts.append(('topic', _slug(persona), persona, 0.80, vis))

    # ── Role family → 'skill' ────────────────────────────────
    role_fam = str(row.get('role_family', ''))
    if role_fam:
        facts.append(('skill', _slug(role_fam), role_fam, 0.75, vis))

    # ── Location → 'location' ────────────────────────────────
    location = str(row.get('location', ''))
    if location:
        facts.append(('location', _slug(location), location, 0.95, vis))

    # ── Objectives from seller/buyer bias ────────────────────
    seller = row.get('seller_bias', 0)
    buyer  = row.get('buyer_bias', 0)
    if isinstance(buyer, float) and buyer > 0.5:
        facts.append(('objective',
                      _slug(f"seeking {pd_desc}"),
                      f"Actively seeking: {pd_desc}", 0.85, vis))
    if isinstance(seller, float) and seller > 0.5:
        facts.append(('offer',
                      _slug(f"providing {pd_desc}"),
                      f"Can provide expertise in: {pd_desc}", 0.85, vis))

    return facts


def step_seed_facts_direct(df, user_ids, doc_ids):
    section("Step 3 — Seed Facts Directly → Postgres + iKG")

    if not HAS_PSYCOPG2:
        print("  ✗  psycopg2 not installed. Run: pip install psycopg2-binary")
        print("     Falling back to LLM extraction via API (slow)...")
        _step_extract_via_api(df, user_ids, doc_ids)
        return

    inserted = 0
    skipped  = 0
    errors   = 0

    try:
        conn = psycopg2.connect(PG_DSN)
        conn.autocommit = False
        cur  = conn.cursor()
    except Exception as e:
        print(f"  ✗  Cannot connect to Postgres: {e}")
        print("     Check PG_DSN at the top of the script.")
        sys.exit(1)

    try:
        for idx, row in df.iterrows():
            name   = row['full_name']
            uid    = user_ids.get(name)
            doc_id = doc_ids.get(name)
            if not uid:
                skipped += 1
                continue

            facts = build_facts_from_row(row, uid)

            for (ftype, canonical, raw, conf, vis) in facts:
                fact_id = str(uuid.uuid4())
                try:
                    cur.execute("""
                        INSERT INTO extracted_facts
                            (fact_id, tenant_id, user_id, fact_type,
                             canonical_value, raw_value, confidence,
                             visibility, source_document_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT DO NOTHING
                    """, (fact_id, TENANT_ID, uid, ftype,
                          canonical, raw, conf, vis, doc_id))
                    inserted += 1
                except Exception as e:
                    errors += 1
                    conn.rollback()

            conn.commit()

            if (idx + 1) % 10 == 0:
                print(f"  →  [{idx+1:3d}/100] {name[:35]:<35}  "
                      f"facts seeded: {len(facts)}")

        cur.close()
        conn.close()

    except Exception as e:
        conn.rollback()
        print(f"  ✗  DB error during fact seeding: {e}")
        sys.exit(1)

    check(errors == 0, f"Facts seeded cleanly  (rows inserted: {inserted}, errors: {errors})")

    # Now push all facts into Memgraph iKG via the API
    section("Step 3b — Sync iKG in Memgraph")
    synced  = 0
    ik_fail = 0
    with httpx.Client(timeout=30) as c:
        for name, uid in user_ids.items():
            r = c.post(f"{BASE_URL}/v1/ikg/upsert",
                       json={"user_id": uid, "tenant_id": TENANT_ID})
            if r.status_code == 200:
                synced += 1
            else:
                ik_fail += 1
            if synced % 10 == 0 and synced > 0:
                print(f"  →  iKG synced: {synced}/{len(user_ids)}")

    check(ik_fail == 0, f"iKG sync complete  (synced: {synced}, failed: {ik_fail})")


def _step_extract_via_api(df, user_ids, doc_ids):
    """Fallback: original LLM extraction path (slow — requires Ollama)."""
    failed = 0
    with httpx.Client(timeout=httpx.Timeout(300, connect=10)) as c:
        for idx, row in df.iterrows():
            name   = row['full_name']
            uid    = user_ids.get(name)
            doc_id = doc_ids.get(name)
            if not doc_id or not uid:
                continue
            for attempt in range(1, 4):
                try:
                    r = c.post(f"{BASE_URL}/v1/ingest/{doc_id}/extract", json={
                        "user_id": uid, "tenant_id": TENANT_ID,
                        "source_type": "cv", "force_reextract": False,
                    })
                    if r.status_code == 200:
                        d = r.json()
                        if (idx + 1) % 10 == 0:
                            print(f"  →  [{idx+1}/100] {name}  facts={d.get('facts_written',0)}")
                        break
                except (httpx.ReadTimeout, httpx.RequestError):
                    time.sleep(2 * attempt)
            else:
                failed += 1
        time.sleep(1.5)
    check(failed == 0, f"LLM extractions done (failed: {failed})")


# ─────────────────────────────────────────────
#  Step 4 — Post intent signals
# ─────────────────────────────────────────────

def step_post_signals(df, user_ids):
    section("Step 4 — Post Intent Signals")
    posted = 0

    with httpx.Client(timeout=30) as c:
        for _, row in df.iterrows():
            name = row['full_name']
            uid  = user_ids[name]

            try:
                needs = json.loads(row['needs_json']) if pd.notna(row['needs_json']) else []
                intent_text = needs[0]['summary'] if needs else row['primary_driver_description']
            except Exception:
                intent_text = str(row['primary_driver_description'])

            r = c.post(f"{BASE_URL}/v1/signals/intent", json={
                "tenant_id":   TENANT_ID,
                "user_id":     uid,
                "signal_type": "intent",
                "payload": {
                    "text":    intent_text,
                    "urgency": "medium",
                    "driver":  row['primary_driver_id'],
                },
            })
            if r.status_code == 200:
                posted += 1

    check(posted > 0, f"Intent signals posted: {posted}/100")


# ─────────────────────────────────────────────
#  Step 5 — Run matchmaking for all users
# ─────────────────────────────────────────────

def step_run_matchmaking(df, user_ids):
    section("Step 5 — Run Matchmaking (100 users)")
    all_matches = {}
    failed = 0

    with httpx.Client(timeout=TIMEOUT_MATCH) as c:
        for idx, row in df.iterrows():
            name = row['full_name']
            uid  = user_ids[name]

            r = c.post(f"{BASE_URL}/v1/matches/generate", json={
                "tenant_id":             TENANT_ID,
                "requesting_user_id":    uid,
                "transaction_types":     [TX_TYPE],
                "max_candidates":        MAX_MATCHES,
                "min_score":             0.01,
                "generate_explanations": False,
            })

            if r.status_code == 200:
                data = r.json()
                matches = data.get("matches", [])
                all_matches[name] = {
                    "uid":     uid,
                    "matches": matches,
                    "count":   len(matches),
                }
            else:
                failed += 1
                all_matches[name] = {"uid": uid, "matches": [], "count": 0}
                info(f"  WARN: Matchmaking failed for {name}: {r.status_code}")

            if (idx + 1) % 10 == 0:
                matched = sum(1 for v in all_matches.values() if v["count"] > 0)
                print(f"  →  Processed {idx + 1}/100  (with matches: {matched})")

    total_matches = sum(v["count"] for v in all_matches.values())
    check(failed == 0, f"All matchmaking calls succeeded (failed: {failed})")
    info(f"Total match pairs generated: {total_matches}")
    return all_matches


# ─────────────────────────────────────────────
#  Step 6 — Save results to Excel
# ─────────────────────────────────────────────

def build_display_name_lookup(df, user_ids):
    return {v: k for k, v in user_ids.items()}

def row_info(df, name):
    row = df[df['full_name'] == name]
    if row.empty:
        return {}
    return row.iloc[0].to_dict()

def step_save_excel(df, user_ids, all_matches, output_path):
    section("Step 6 — Save Results to Excel")

    uid_to_name = build_display_name_lookup(df, user_ids)
    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────
    header_fill    = PatternFill("solid", fgColor="1F4E79")
    alt_fill       = PatternFill("solid", fgColor="EBF3FA")
    score_hi_fill  = PatternFill("solid", fgColor="C6EFCE")
    score_med_fill = PatternFill("solid", fgColor="FFEB9C")
    score_lo_fill  = PatternFill("solid", fgColor="FFC7CE")
    center_align   = Alignment(horizontal="center", vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin",  color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    def hdr(cell):
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    def std(cell, alt=False, align="center"):
        cell.font      = Font(name="Arial", size=9)
        cell.fill      = alt_fill if alt else PatternFill()
        cell.alignment = center_align if align == "center" else left_align
        cell.border    = thin_border

    def sfill(val):
        if val is None: return PatternFill()
        if val >= 0.6:  return score_hi_fill
        if val >= 0.3:  return score_med_fill
        return score_lo_fill

    # ── Sheet 1: Match Results ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Match Results"

    h1 = [
        "Requester Name", "Requester Role", "Requester Org", "Requester Location",
        "Requester Persona", "Requester Primary Driver",
        "Rank", "Matched Name", "Matched Role", "Matched Org", "Matched Location",
        "Matched Persona", "Matched Primary Driver",
        "Score", "Relevance", "Complementarity", "Timing", "Proximity",
        "Evidence Strength", "Outcome Likelihood", "Novelty",
        "Privacy Risk", "Interaction Friction",
        "Match ID", "Transaction Type",
    ]
    for col, h in enumerate(h1, 1):
        hdr(ws1.cell(row=1, column=col, value=h))
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    row_num = 2
    for idx, (name, data) in enumerate(all_matches.items()):
        req  = row_info(df, name)
        mats = data.get("matches", [])
        alt  = (idx % 2 == 0)

        if not mats:
            vals = [name, req.get("current_role",""), req.get("organisation_name",""),
                    req.get("location",""), req.get("persona_cluster",""),
                    req.get("primary_driver_description",""),
                    "—", "No matches found", "", "", "", "", "",
                    *([None]*9), "", TX_TYPE]
            for col, val in enumerate(vals, 1):
                c = ws1.cell(row=row_num, column=col, value=val)
                std(c, alt=alt, align="left" if col <= 6 else "center")
            row_num += 1
            continue

        for rank, m in enumerate(mats, 1):
            matched_uid  = m.get("person_b", "")
            matched_name = uid_to_name.get(matched_uid, matched_uid)
            mat          = row_info(df, matched_name)
            bs           = m.get("score_breakdown") or {}
            score        = m.get("score")

            def g(k): return bs.get(k)

            vals = [
                name, req.get("current_role",""), req.get("organisation_name",""),
                req.get("location",""), req.get("persona_cluster",""),
                req.get("primary_driver_description",""),
                rank,
                matched_name, mat.get("current_role",""), mat.get("organisation_name",""),
                mat.get("location",""), mat.get("persona_cluster",""),
                mat.get("primary_driver_description",""),
                score,
                g("relevance"), g("complementarity"), g("timing"), g("proximity"),
                g("evidence_strength"), g("outcome_likelihood"), g("novelty"),
                g("privacy_risk"), g("interaction_friction"),
                m.get("match_id",""), m.get("transaction_type", TX_TYPE),
            ]

            for col, val in enumerate(vals, 1):
                c = ws1.cell(row=row_num, column=col, value=val)
                if 14 <= col <= 22:
                    c.font = Font(name="Arial", size=9)
                    c.fill = sfill(val)
                    c.alignment = center_align
                    c.border = thin_border
                    if isinstance(val, float):
                        c.number_format = "0.000"
                else:
                    align = "left" if col in (1,2,3,6,8,9,10,13) else "center"
                    std(c, alt=alt, align=align)
            row_num += 1

    widths_1 = [22,28,22,14,22,45, 7, 22,28,22,14,22,45, 8,8,8,8,8,8,8,8,8,8, 36,22]
    for i, w in enumerate(widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: User Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("User Summary")
    h2 = ["Full Name","Current Role","Organisation","Archetype","Location","Seniority",
          "Role Family","Persona Cluster","Market Regime","Primary Driver","Secondary Driver",
          "Seller Bias","Buyer Bias","Matches Found","Top Match","Top Score","User ID"]
    for col, h in enumerate(h2, 1):
        hdr(ws2.cell(row=1, column=col, value=h))
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 30

    for idx, (_, row) in enumerate(df.iterrows()):
        name    = row['full_name']
        data    = all_matches.get(name, {})
        matches = data.get("matches", [])
        uid     = user_ids.get(name, "")
        top_name, top_score = "", None
        if matches:
            t = matches[0]
            top_name  = uid_to_name.get(t.get("person_b",""), "")
            top_score = t.get("score")

        alt  = (idx % 2 == 0)
        vals = [name, row.get("current_role",""), row.get("organisation_name",""),
                row.get("organisation_archetype",""), row.get("location",""),
                row.get("seniority",""), row.get("role_family",""),
                row.get("persona_cluster",""), row.get("market_regime",""),
                row.get("primary_driver_description",""), row.get("secondary_driver_description",""),
                row.get("seller_bias"), row.get("buyer_bias"),
                len(matches), top_name, top_score, uid]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=idx+2, column=col, value=val)
            if col == 16 and top_score is not None:
                c.font = Font(name="Arial", size=9)
                c.fill = sfill(val)
                c.alignment = center_align
                c.border = thin_border
                c.number_format = "0.000"
            else:
                align = "left" if col in (1,2,3,10,11,15) else "center"
                std(c, alt=alt, align=align)

    for i, w in enumerate([22,30,22,22,14,12,16,24,26,50,50,10,10,10,22,10,36], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Analytics ───────────────────────────────────────
    ws3 = wb.create_sheet("Analytics")
    total_users   = len(df)
    users_matched = sum(1 for v in all_matches.values() if v["count"] > 0)
    total_pairs   = sum(v["count"] for v in all_matches.values())
    all_scores    = [m.get("score",0) for v in all_matches.values()
                     for m in v.get("matches",[]) if m.get("score") is not None]
    avg_score = sum(all_scores)/len(all_scores) if all_scores else 0
    high   = sum(1 for s in all_scores if s >= 0.6)
    medium = sum(1 for s in all_scores if 0.3 <= s < 0.6)
    low    = sum(1 for s in all_scores if s < 0.3)

    rows_analytics = [
        ("TENANT OVERVIEW", None),
        ("Tenant ID",           TENANT_ID),
        ("Tenant Name",         TENANT_NAME),
        ("Run Date",            datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Transaction Type",    TX_TYPE),
        ("Max Matches/User",    MAX_MATCHES),
        ("", None),
        ("USER STATS", None),
        ("Total Users",         total_users),
        ("Users With Matches",  users_matched),
        ("Users Without Matches", total_users - users_matched),
        ("Match Coverage",      f"{users_matched/total_users*100:.1f}%"),
        ("", None),
        ("MATCH STATS", None),
        ("Total Match Pairs",   total_pairs),
        ("Avg Matches/User",    f"{total_pairs/total_users:.1f}"),
        ("Average Score",       f"{avg_score:.4f}"),
        ("Max Score",           max(all_scores) if all_scores else 0),
        ("Min Score",           min(all_scores) if all_scores else 0),
        ("", None),
        ("SCORE BANDS", None),
        ("High (≥ 0.60)",       high),
        ("Medium (0.30–0.59)",  medium),
        ("Low (< 0.30)",        low),
    ]

    for rn, (label, val) in enumerate(rows_analytics, 1):
        cl = ws3.cell(row=rn, column=1, value=label)
        cv = ws3.cell(row=rn, column=2, value=val)
        cl.border = thin_border
        cv.border = thin_border
        if val is None:
            cl.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cl.fill = header_fill
            cl.alignment = left_align
        else:
            cl.font = Font(name="Arial", bold=True, size=10)
            cl.alignment = left_align
            cv.font = Font(name="Arial", size=10)
            cv.alignment = center_align

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 42

    # ── Sheet 4: Driver Cross-Match ───────────────────────────────
    ws4 = wb.create_sheet("Driver Cross-Match")
    driver_pairs = {}
    for name, data in all_matches.items():
        req_driver = row_info(df, name).get("primary_driver_id", "unknown")
        for m in data.get("matches", []):
            mat_name   = uid_to_name.get(m.get("person_b",""), "")
            mat_driver = row_info(df, mat_name).get("primary_driver_id", "unknown")
            k = (req_driver, mat_driver)
            driver_pairs[k] = driver_pairs.get(k, 0) + 1

    for col, h in enumerate(["Requester Driver", "Matched Driver", "Match Count"], 1):
        hdr(ws4.cell(row=1, column=col, value=h))
    ws4.row_dimensions[1].height = 25

    for rn, ((d1, d2), cnt) in enumerate(sorted(driver_pairs.items(), key=lambda x: -x[1]), 2):
        alt = (rn % 2 == 0)
        for col, val in enumerate([d1, d2, cnt], 1):
            std(ws4.cell(row=rn, column=col, value=val), alt=alt,
                align="left" if col < 3 else "center")

    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 42
    ws4.column_dimensions["C"].width = 14

    wb.save(output_path)
    check(True, f"Results saved → {output_path}")
    info("Sheets: Match Results | User Summary | Analytics | Driver Cross-Match")


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    print("\n🚀  Delllo RAIN3.0 — CIL 100-User Matchmaking Test")
    print(f"    API:    {BASE_URL}")
    print(f"    Tenant: {TENANT_ID}")
    print(f"    Input:  {INPUT_FILE}")
    print(f"    Output: {OUTPUT_FILE}")
    if HAS_PSYCOPG2:
        print("    Mode:   FAST (direct Postgres fact seeding — no Ollama needed)")
    else:
        print("    Mode:   SLOW (LLM extraction via Ollama)")
        print("    TIP:    pip install psycopg2-binary  to use fast mode\n")
    print()

    df = pd.read_excel(INPUT_FILE)
    info(f"Loaded {len(df)} users from {INPUT_FILE}")
    check(len(df) == 100, "Dataset has 100 users")

    t0 = time.time()

    step_setup_tenant()
    user_ids    = step_create_users(df)
    doc_ids     = step_ingest(df, user_ids)
    step_seed_facts_direct(df, user_ids, doc_ids)   # fast path — no Ollama
    step_post_signals(df, user_ids)
    all_matches = step_run_matchmaking(df, user_ids)
    step_save_excel(df, user_ids, all_matches, OUTPUT_FILE)

    elapsed = time.time() - t0
    total_m = sum(v["count"] for v in all_matches.values())
    passed  = sum(1 for ok, _ in results_log if ok)
    failed  = sum(1 for ok, _ in results_log if not ok)

    section("Final Summary")
    info(f"Total time:           {elapsed:.0f}s")
    info(f"Users processed:      100")
    info(f"Match pairs found:    {total_m}")
    info(f"Checks passed/failed: {passed}/{failed}")
    print()
    if failed == 0:
        print("  ✅  All checks passed — results saved to matchmaking_results.xlsx")
    else:
        print(f"  ⚠️   {failed} checks failed — see output above")
    print()


if __name__ == "__main__":
    main()