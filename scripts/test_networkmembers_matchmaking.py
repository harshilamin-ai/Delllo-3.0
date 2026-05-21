#!/usr/bin/env python3
"""
Delllo RAIN3.0 — NetworkMembersA Matchmaking Test
══════════════════════════════════════════════════════════════════
Reads NetworkMembersA_updated_20260517_162617.json (103 users),
creates a fresh tenant, seeds facts directly into Postgres,
syncs iKG, posts intent signals, runs matchmaking for every user
(top-3 candidates), and saves results to Excel.

Steps:
  0  Health check
  1  Wipe + create fresh tenant
  2  Create 103 users
  3  Ingest CV text
  4  Seed facts directly → Postgres (no Ollama needed)
  5  Sync iKG → Memgraph
  6  Post intent signals
  7  Run matchmaking for every user → top-3
  8  Save to Excel + console summary
"""

import sys, json, time, uuid, re
import httpx
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import psycopg2, psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False

# ─── Config ───────────────────────────────────────────────────
BASE_URL    = "http://localhost:8000"
TENANT_ID   = "aaaa0001-0000-0000-0000-000000000001"
TENANT_NAME = "NetworkMembersA Test"
TENANT_SLUG = "netmem-a-test"
PG_DSN      = "dbname=delllo_db user=delllo password=delllo_secret host=localhost port=5432"

SCRIPT_DIR  = Path(__file__).parent.resolve()
INPUT_FILE  = SCRIPT_DIR / "NetworkMembersA_updated_20260517_162617.json"
OUTPUT_FILE = SCRIPT_DIR.parent / "NetworkMembersA_Matchmaking_Results.xlsx"
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

TX_TYPE     = "knowledge_transfer"
MAX_MATCHES = 3
TIMEOUT_INGEST = httpx.Timeout(3000.0, connect=30.0)
TIMEOUT_MATCH  = httpx.Timeout(3000.0, connect=30.0)

# ─── Logging ──────────────────────────────────────────────────
log = []

def section(t):
    print(f"\n{'═'*64}\n  {t}\n{'═'*64}")

def info(m):
    print(f"     {m}")

def check(ok, msg, fatal=False):
    print(f"  {'✓' if ok else '✗'}  {msg}")
    log.append((ok, msg))
    if not ok and fatal:
        print("  ❌ Fatal — stopping.")
        sys.exit(1)

# ─── ID helper ────────────────────────────────────────────────
_NS = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

def make_uid(name):
    return str(uuid.uuid5(_NS, f"{TENANT_ID}:{name}"))

# ─── Load dataset ─────────────────────────────────────────────
def load_data():
    with open(INPUT_FILE) as f:
        raw = json.load(f)
    records = raw["sheets"]["Sheet1"]["records"]
    info(f"Loaded {len(records)} users from {INPUT_FILE.name}")
    return records

# ─── Build CV text ────────────────────────────────────────────
def _parse(val):
    if not val:
        return []
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        try:
            return json.loads(val)
        except Exception:
            return []
    return []

def _slug(t):
    return re.sub(r'[^a-z0-9]+', '_', str(t).lower()).strip('_')[:120]

def build_cv(row):
    lines = [
        f"{row['full_name']} — {row['current_role']}, {row['organisation_name']}",
        "",
        "IDENTITY",
        f"Organisation: {row['organisation_name']} ({row['organisation_archetype']})",
        f"Location: {row['location']}",
        f"Seniority: {row['seniority']}",
        f"Role Family: {row['role_family']}",
        f"Persona: {row['persona_cluster']}",
        f"Market Regime: {row['market_regime']}",
        "",
    ]

    hist = row.get('previous_employment_and_education_history_json') or {}
    if isinstance(hist, str):
        try:
            hist = json.loads(hist)
        except Exception:
            hist = {}
    emp = hist.get("employment_history", [])
    if emp:
        lines.append("EXPERIENCE")
        for e in emp:
            lines.append(f"- {e.get('role','?')} at {e.get('organisation','?')} ({e.get('period','?')})")
        lines.append("")

    lines += [
        "MARKET DRIVERS",
        f"Primary Driver: {row['primary_driver_description']}",
        f"Secondary Driver: {row['secondary_driver_description']}",
        "",
    ]

    for o in _parse(row.get('offers_json')):
        lines.append(f"OFFER: [{o.get('offer_type','?')}] {o.get('summary','')}")
    lines.append("")

    for n in _parse(row.get('needs_json')):
        lines.append(f"NEED: [{n.get('need_type','?')}] {n.get('summary','')}")
    lines.append("")

    for i in _parse(row.get('insights_json')):
        lines.append(f"INSIGHT: [{i.get('insight_type','?')}] {i.get('summary','')}")
    lines.append("")

    for s in _parse(row.get('solutions_json')):
        lines.append(f"SOLUTION: {s.get('summary','')} (prereqs: {', '.join(s.get('prerequisites',[]))})")
    lines.append("")

    return "\n".join(lines)

# ─── Build facts ──────────────────────────────────────────────
def build_facts(row):
    facts = []
    vis   = "match_engine_only"

    for o in _parse(row.get('offers_json')):
        raw = o.get('summary', '')
        if raw:
            facts.append(('offer', _slug(raw), raw, 0.90, vis))

    for n in _parse(row.get('needs_json')):
        raw = n.get('summary', '')
        if raw:
            facts.append(('need', _slug(raw), raw, 0.90, vis))

    for s in _parse(row.get('solutions_json')):
        raw = s.get('summary', '')
        if raw:
            facts.append(('skill', _slug(raw), raw, 0.85, vis))

    for i in _parse(row.get('insights_json')):
        raw = i.get('summary', '')
        if raw:
            facts.append(('topic', _slug(raw), raw, 0.80, vis))

    pd_desc = str(row.get('primary_driver_description', ''))
    if pd_desc:
        facts.append(('domain', _slug(row.get('primary_driver_id', pd_desc)), pd_desc, 0.95, vis))

    sd_desc = str(row.get('secondary_driver_description', ''))
    if sd_desc and sd_desc != pd_desc:
        facts.append(('domain', _slug(row.get('secondary_driver_id', sd_desc)), sd_desc, 0.85, vis))

    persona = str(row.get('persona_cluster', ''))
    if persona:
        facts.append(('topic', _slug(persona), persona, 0.80, vis))

    role_fam = str(row.get('role_family', ''))
    if role_fam:
        facts.append(('skill', _slug(role_fam), role_fam, 0.75, vis))

    loc = str(row.get('location', ''))
    if loc:
        facts.append(('location', _slug(loc), loc, 0.95, vis))

    buyer  = float(row.get('buyer_bias', 0) or 0)
    seller = float(row.get('seller_bias', 0) or 0)
    if buyer > 0.5:
        facts.append(('objective', _slug(f"seeking_{pd_desc}")[:80],
                      f"Actively seeking: {pd_desc}", 0.85, vis))
    if seller > 0.5:
        facts.append(('offer', _slug(f"providing_{pd_desc}")[:80],
                      f"Can provide expertise in: {pd_desc}", 0.85, vis))

    seen, out = set(), []
    for f in facts:
        key = (f[0], f[1])
        if key not in seen:
            seen.add(key)
            out.append(f)
    return out

# ─── Step 0: Health ───────────────────────────────────────────
def step_health():
    section("Step 0 — Health Check")
    with httpx.Client(timeout=30) as c:
        r = c.get(f"{BASE_URL}/health")
        check(r.status_code == 200, "API reachable", fatal=True)
        r2 = c.get(f"{BASE_URL}/health/stack")
        if r2.status_code == 200:
            for svc, d in r2.json().get("services", {}).items():
                ok = d.get("status") == "ok"
                print(f"  {'✓' if ok else '⚠'}  {svc}: {str(d.get('detail',''))[:70]}")

# ─── Step 1: Tenant ───────────────────────────────────────────
def step_tenant():
    section("Step 1 — Create / Reset Tenant")
    with httpx.Client(timeout=30) as c:
        r = c.post(f"{BASE_URL}/v1/admin/wipe",
                   json={"tenant_id": TENANT_ID, "confirm": True})
        if r.status_code == 200:
            info(f"Wiped prior data: {r.json().get('status')}")
        else:
            info("No prior data to wipe (first run)")

        r = c.post(f"{BASE_URL}/v1/tenants", json={
            "tenant_id": TENANT_ID,
            "name":      TENANT_NAME,
            "slug":      TENANT_SLUG,
        })
        if r.status_code in (200, 201):
            info(f"Tenant created: {TENANT_ID[:8]}… slug={TENANT_SLUG}")
        elif r.status_code == 409:
            info(f"Tenant already exists (slug conflict — wipe may not have cleared it)")
        else:
            info(f"Tenant HTTP {r.status_code}: {r.text[:120]}")

# ─── Step 2: Create users ─────────────────────────────────────
def step_create_users(records):
    section("Step 2 — Create Users")
    user_ids = {}
    with httpx.Client(timeout=30) as c:
        for i, row in enumerate(records):
            name  = row['full_name']
            uid   = make_uid(name)
            email = re.sub(r'[^a-z0-9]', '.', name.lower()) + "@netmem-test.com"
            r = c.post(f"{BASE_URL}/v1/users", json={
                "user_id":      uid,
                "tenant_id":    TENANT_ID,
                "display_name": name,
                "email":        email,
                "headline":     f"{row['current_role']} at {row['organisation_name']}",
                "role":         "member",
                "status":       "active",
            })
            ok = r.status_code in (200, 201)
            user_ids[name] = uid
            if (i + 1) % 25 == 0:
                check(ok, f"Users created: {i+1}/{len(records)}")
    check(len(user_ids) == len(records), f"All {len(records)} users registered")
    return user_ids

# ─── Step 3: Ingest CV text ───────────────────────────────────
def step_ingest(records, user_ids):
    section("Step 3 — Ingest CV Profiles")
    doc_ids = {}
    failed  = 0
    with httpx.Client(timeout=TIMEOUT_INGEST) as c:
        for i, row in enumerate(records):
            name = row['full_name']
            uid  = user_ids[name]
            r = c.post(f"{BASE_URL}/v1/ingest/text", data={
                "tenant_id":   TENANT_ID,
                "user_id":     uid,
                "content":     build_cv(row),
                "source_type": "cv",
                "filename":    f"{name.replace(' ','_')}_profile.txt",
                "embed":       "true",
            })
            if r.status_code == 200:
                doc_ids[name] = r.json()["document_id"]
            else:
                failed += 1
                info(f"  WARN ingest failed {name}: {r.status_code}")
            if (i + 1) % 25 == 0:
                print(f"  →  Ingested {i+1}/{len(records)}")
    check(failed == 0, f"All ingestions succeeded (failed: {failed})")
    info(f"Documents created: {len(doc_ids)}")
    return doc_ids

# ─── Step 4: Seed facts → Postgres ────────────────────────────
def step_seed_facts(records, user_ids, doc_ids):
    section("Step 4 — Seed Facts → Postgres")
    if not HAS_PG:
        print("  ✗  psycopg2 not installed — pip install psycopg2-binary")
        sys.exit(1)

    inserted = errors = 0
    try:
        conn = psycopg2.connect(PG_DSN)
        conn.autocommit = False
        cur  = conn.cursor()
    except Exception as e:
        print(f"  ✗  Cannot connect to Postgres: {e}")
        sys.exit(1)

    try:
        for i, row in enumerate(records):
            name   = row['full_name']
            uid    = user_ids.get(name)
            doc_id = doc_ids.get(name)
            if not uid:
                continue
            for (ftype, canonical, raw, conf, vis) in build_facts(row):
                fid = str(uuid.uuid4())
                try:
                    cur.execute("""
                        INSERT INTO extracted_facts
                            (fact_id, tenant_id, user_id, fact_type,
                             canonical_value, raw_value, confidence,
                             visibility, source_document_id)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT DO NOTHING
                    """, (fid, TENANT_ID, uid, ftype, canonical, raw, conf, vis, doc_id))
                    inserted += 1
                except Exception as e:
                    errors += 1
                    conn.rollback()
            conn.commit()
            if (i + 1) % 25 == 0:
                print(f"  →  [{i+1:3d}/{len(records)}] facts seeded…")
        cur.close()
        conn.close()
    except Exception as e:
        conn.rollback()
        print(f"  ✗  DB error: {e}")
        sys.exit(1)

    check(errors == 0, f"Facts seeded (inserted: {inserted}, errors: {errors})")

# ─── Step 5: Sync iKG ─────────────────────────────────────────
def step_sync_ikg(user_ids):
    section("Step 5 — Sync iKG → Memgraph")
    synced = fails = 0
    with httpx.Client(timeout=httpx.Timeout(60.0)) as c:
        for i, (name, uid) in enumerate(user_ids.items()):
            r = c.post(f"{BASE_URL}/v1/ikg/upsert",
                       json={"user_id": uid, "tenant_id": TENANT_ID})
            if r.status_code == 200:
                synced += 1
            else:
                fails += 1
            if (i + 1) % 25 == 0:
                print(f"  →  iKG synced: {i+1}/{len(user_ids)}")
    check(fails == 0, f"iKG sync complete (synced: {synced}, failed: {fails})")

# ─── Step 6: Post intent signals ──────────────────────────────
def step_signals(records, user_ids):
    section("Step 6 — Post Intent Signals")
    posted = 0
    with httpx.Client(timeout=30) as c:
        for row in records:
            name  = row['full_name']
            uid   = user_ids[name]
            needs = _parse(row.get('needs_json'))
            intent = needs[0]['summary'] if needs else row['primary_driver_description']
            r = c.post(f"{BASE_URL}/v1/signals/intent", json={
                "tenant_id":   TENANT_ID,
                "user_id":     uid,
                "signal_type": "intent",
                "payload": {
                    "text":    intent,
                    "urgency": "medium",
                    "driver":  row.get('primary_driver_id', ''),
                },
            })
            if r.status_code == 200:
                posted += 1
    check(posted > 0, f"Intent signals posted: {posted}/{len(records)}")

# ─── Step 7: Run matchmaking ───────────────────────────────────
def step_matchmaking(records, user_ids):
    section("Step 7 — Run Matchmaking (top-3 per user)")
    all_matches = {}
    failed = 0
    with httpx.Client(timeout=TIMEOUT_MATCH) as c:
        for i, row in enumerate(records):
            name = row['full_name']
            uid  = user_ids[name]
            r = c.post(f"{BASE_URL}/v1/matches/generate", json={
                "tenant_id":             TENANT_ID,
                "requesting_user_id":    uid,
                "transaction_types":     [TX_TYPE],
                "max_candidates":        MAX_MATCHES,
                "min_score":             0.01,
                "generate_explanations": False,
            })
            if r.status_code == 200:
                matches = r.json().get("matches", [])
                all_matches[name] = {"uid": uid, "matches": matches, "count": len(matches)}
            else:
                failed += 1
                all_matches[name] = {"uid": uid, "matches": [], "count": 0}
                info(f"  WARN failed {name}: {r.status_code} {r.text[:100]}")
            if (i + 1) % 20 == 0:
                matched = sum(1 for v in all_matches.values() if v["count"] > 0)
                print(f"  →  Processed {i+1}/{len(records)} (with matches: {matched})")

    total = sum(v["count"] for v in all_matches.values())
    check(failed == 0, f"All matchmaking calls succeeded (failed: {failed})")
    info(f"Total match pairs: {total}")
    return all_matches

# ─── Step 8: Save Excel ────────────────────────────────────────
def step_save_excel(records, user_ids, all_matches):
    section("Step 8 — Save Results to Excel")

    uid_to_name = {v: k for k, v in user_ids.items()}
    rec_by_name = {r['full_name']: r for r in records}

    wb = Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="EBF3FA")
    hi_fill  = PatternFill("solid", fgColor="C6EFCE")
    med_fill = PatternFill("solid", fgColor="FFEB9C")
    lo_fill  = PatternFill("solid", fgColor="FFC7CE")
    c_align  = Alignment(horizontal="center", vertical="center")
    l_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def hdr(cell):
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill; cell.alignment = c_align; cell.border = thin

    def std(cell, alt=False, align="center"):
        cell.font = Font(name="Arial", size=9)
        cell.fill = alt_fill if alt else PatternFill()
        cell.alignment = c_align if align == "center" else l_align
        cell.border = thin

    def sfill(v):
        if v is None: return PatternFill()
        if v >= 0.6:  return hi_fill
        if v >= 0.3:  return med_fill
        return lo_fill

    # ── Sheet 1: Match Results ──────────────────────────────────
    ws1 = wb.active; ws1.title = "Match Results"
    h1 = [
        "Requester","Requester Role","Requester Org","Location","Persona","Primary Driver",
        "Rank","Match Name","Match Role","Match Org","Match Location","Match Persona","Match Driver",
        "Score","Relevance","Complementarity","Timing","Proximity",
        "Evidence","Outcome","Novelty","Privacy","Friction",
        "Match ID","Transaction Type",
    ]
    for col, h in enumerate(h1, 1):
        hdr(ws1.cell(row=1, column=col, value=h))
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    row_num = 2
    for idx, (name, data) in enumerate(all_matches.items()):
        req  = rec_by_name.get(name, {})
        mats = data.get("matches", [])
        alt  = (idx % 2 == 0)
        if not mats:
            vals = [name, req.get("current_role",""), req.get("organisation_name",""),
                    req.get("location",""), req.get("persona_cluster",""),
                    req.get("primary_driver_description",""),
                    "—", "No matches found", "", "", "", "", "",
                    *([None]*9), "", TX_TYPE]
            for col, val in enumerate(vals, 1):
                std(ws1.cell(row=row_num, column=col, value=val), alt=alt,
                    align="left" if col <= 6 else "center")
            row_num += 1
            continue

        for rank, m in enumerate(mats, 1):
            mid_uid  = m.get("person_b", "")
            mat_name = uid_to_name.get(mid_uid, mid_uid[:8] + "…")
            mat      = rec_by_name.get(mat_name, {})
            bs       = m.get("score_breakdown") or {}
            score    = m.get("score")
            vals = [
                name, req.get("current_role",""), req.get("organisation_name",""),
                req.get("location",""), req.get("persona_cluster",""),
                req.get("primary_driver_description",""),
                rank, mat_name, mat.get("current_role",""), mat.get("organisation_name",""),
                mat.get("location",""), mat.get("persona_cluster",""),
                mat.get("primary_driver_description",""),
                score,
                bs.get("relevance"), bs.get("complementarity"), bs.get("timing"),
                bs.get("proximity"), bs.get("evidence_strength"),
                bs.get("outcome_likelihood"), bs.get("novelty"),
                bs.get("privacy_risk"), bs.get("interaction_friction"),
                m.get("match_id",""), TX_TYPE,
            ]
            for col, val in enumerate(vals, 1):
                c = ws1.cell(row=row_num, column=col, value=val)
                if 14 <= col <= 22:
                    c.font = Font(name="Arial", size=9); c.fill = sfill(val)
                    c.alignment = c_align; c.border = thin
                    if isinstance(val, float): c.number_format = "0.000"
                else:
                    std(c, alt=alt,
                        align="left" if col in (1,2,3,6,8,9,10,13) else "center")
            row_num += 1

    for i, w in enumerate([22,30,24,12,20,50, 6,
                             22,30,24,12,20,50,
                             8,7,7,7,7,7,7,7,7,7,36,18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Top-3 Summary (clean) ─────────────────────────
    ws2 = wb.create_sheet("Top-3 Summary")
    h2  = ["Requester","Role","Org","Location",
           "#1 Name","#1 Score","#1 Org","#1 Driver",
           "#2 Name","#2 Score","#2 Org","#2 Driver",
           "#3 Name","#3 Score","#3 Org","#3 Driver"]
    for col, h in enumerate(h2, 1):
        hdr(ws2.cell(row=1, column=col, value=h))
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 28

    for idx, (name, data) in enumerate(all_matches.items()):
        req  = rec_by_name.get(name, {})
        mats = data.get("matches", [])
        alt  = (idx % 2 == 0)
        rv   = [name, req.get("current_role",""),
                req.get("organisation_name",""), req.get("location","")]
        for ri in range(3):
            if ri < len(mats):
                m        = mats[ri]
                mid_uid  = m.get("person_b","")
                mat_name = uid_to_name.get(mid_uid, mid_uid[:8]+"…")
                mat      = rec_by_name.get(mat_name, {})
                rv += [mat_name, m.get("score"),
                       mat.get("organisation_name",""),
                       str(mat.get("primary_driver_description",""))[:60]]
            else:
                rv += ["—", None, "", ""]
        for col, val in enumerate(rv, 1):
            c = ws2.cell(row=idx+2, column=col, value=val)
            if col in (6, 10, 14) and val is not None:
                c.font = Font(name="Arial", size=9); c.fill = sfill(val)
                c.alignment = c_align; c.border = thin; c.number_format = "0.000"
            else:
                std(c, alt=alt,
                    align="left" if col in (1,2,3,5,7,8,9,11,12,13,15,16) else "center")

    for i, w in enumerate([22,30,24,12,
                             22,8,24,45, 22,8,24,45, 22,8,24,45], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Analytics ──────────────────────────────────────
    ws3 = wb.create_sheet("Analytics")
    all_scores = [m.get("score",0) for v in all_matches.values()
                  for m in v.get("matches",[]) if m.get("score") is not None]
    avg_score = sum(all_scores)/len(all_scores) if all_scores else 0
    users_matched = sum(1 for v in all_matches.values() if v["count"] > 0)
    total_pairs   = sum(v["count"] for v in all_matches.values())
    high   = sum(1 for s in all_scores if s >= 0.6)
    medium = sum(1 for s in all_scores if 0.3 <= s < 0.6)
    low    = sum(1 for s in all_scores if s < 0.3)

    a_rows = [
        ("TENANT OVERVIEW", None),
        ("Tenant ID",        TENANT_ID),
        ("Tenant Name",      TENANT_NAME),
        ("Run Date",         datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Transaction Type", TX_TYPE),
        ("Top-N per user",   MAX_MATCHES),
        ("", None),
        ("USER STATS", None),
        ("Total Users",      len(records)),
        ("Users With Matches", users_matched),
        ("Users Without Matches", len(records) - users_matched),
        ("Coverage",         f"{users_matched/len(records)*100:.1f}%"),
        ("", None),
        ("MATCH STATS", None),
        ("Total Match Pairs", total_pairs),
        ("Avg Matches/User",  f"{total_pairs/len(records):.2f}"),
        ("Average Score",     f"{avg_score:.4f}"),
        ("Max Score",         max(all_scores) if all_scores else 0),
        ("Min Score",         min(all_scores) if all_scores else 0),
        ("", None),
        ("SCORE BANDS", None),
        ("High (≥ 0.60)",    high),
        ("Medium (0.30–0.59)", medium),
        ("Low (< 0.30)",     low),
    ]
    for rn, (label, val) in enumerate(a_rows, 1):
        cl = ws3.cell(row=rn, column=1, value=label)
        cv = ws3.cell(row=rn, column=2, value=val)
        cl.border = thin; cv.border = thin
        if val is None:
            cl.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cl.fill = hdr_fill; cl.alignment = l_align
        else:
            cl.font = Font(name="Arial", bold=True, size=10); cl.alignment = l_align
            cv.font = Font(name="Arial", size=10);            cv.alignment = c_align
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 42

    wb.save(OUTPUT_FILE)
    check(True, f"Results saved → {OUTPUT_FILE}")
    info("Sheets: Match Results | Top-3 Summary | Analytics")

# ─── Console summary ──────────────────────────────────────────
def print_summary(records, user_ids, all_matches):
    section("Console — Top-3 Matches (first 25 users with matches)")
    uid_to_name = {v: k for k, v in user_ids.items()}
    rec_by_name = {r['full_name']: r for r in records}
    shown = 0
    for name, data in all_matches.items():
        mats = data.get("matches", [])
        if not mats:
            continue
        req = rec_by_name.get(name, {})
        print(f"\n  👤 {name}")
        print(f"     {req.get('current_role','')} @ {req.get('organisation_name','')}  [{req.get('location','')}]")
        for rank, m in enumerate(mats, 1):
            mid_uid  = m.get("person_b", "")
            mat_name = uid_to_name.get(mid_uid, mid_uid[:8]+"…")
            mat      = rec_by_name.get(mat_name, {})
            score    = m.get("score", 0)
            bs       = m.get("score_breakdown") or {}
            bar      = "🟢" if score >= 0.6 else "🟡" if score >= 0.3 else "🔴"
            print(f"     {bar} #{rank}  {mat_name:<30} score={score:.3f}"
                  f"  rel={bs.get('relevance',0):.2f}"
                  f"  comp={bs.get('complementarity',0):.2f}"
                  f"  evid={bs.get('evidence_strength',0):.2f}")
            if mat:
                print(f"          {mat.get('current_role','')} @ {mat.get('organisation_name','')}  [{mat.get('location','')}]")
        shown += 1
        if shown >= 25:
            remaining = sum(1 for v in all_matches.values() if v["count"] > 0) - shown
            if remaining > 0:
                print(f"\n  … {remaining} more users in the Excel output")
            break

# ─── Main ─────────────────────────────────────────────────────
def main():
    print(f"\n🚀  Delllo RAIN3.0 — NetworkMembersA Matchmaking Test")
    print(f"    API:    {BASE_URL}")
    print(f"    Tenant: {TENANT_ID}")
    print(f"    Input:  {INPUT_FILE.name}")
    print(f"    Output: {OUTPUT_FILE}")
    print(f"    Mode:   {'FAST (direct Postgres)' if HAS_PG else 'SLOW (no psycopg2)'}\n")

    records = load_data()
    t0 = time.time()

    step_health()
    step_tenant()
    user_ids    = step_create_users(records)
    doc_ids     = step_ingest(records, user_ids)
    step_seed_facts(records, user_ids, doc_ids)
    step_sync_ikg(user_ids)
    step_signals(records, user_ids)
    all_matches = step_matchmaking(records, user_ids)
    step_save_excel(records, user_ids, all_matches)
    print_summary(records, user_ids, all_matches)

    elapsed     = time.time() - t0
    total_pairs = sum(v["count"] for v in all_matches.values())
    passed      = sum(1 for ok, _ in log if ok)
    failed      = sum(1 for ok, _ in log if not ok)

    section("Final Summary")
    info(f"Total time:      {elapsed:.0f}s")
    info(f"Users processed: {len(records)}")
    info(f"Match pairs:     {total_pairs}")
    info(f"Checks:          {passed} passed / {failed} failed")
    print()
    if failed == 0:
        print("  ✅  All checks passed — results saved to Excel")
    else:
        print(f"  ⚠️   {failed} check(s) failed — see output above")
    print()

if __name__ == "__main__":
    main()