#!/usr/bin/env python3
"""
Delllo RAIN3.0 — CDL Network Matchmaking Test
═══════════════════════════════════════════════════════════════════
Reads NetworkMembersA_updated JSON, creates a fresh CDL tenant,
seeds all users + facts directly into Postgres (no Ollama),
runs matchmaking for every user (top 3 candidates), and saves
full results + expected-match validation to Excel.

Usage:
  pip install httpx openpyxl psycopg2-binary
  python scripts/test_cdl_matchmaking.py

Place NetworkMembersA_updated_*.json in the Rain 3.0 root dir.
═══════════════════════════════════════════════════════════════════
"""

import json
import re
import sys
import time
import uuid as _uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import psycopg2
    HAS_PG = True
except ImportError:
    HAS_PG = False

# ─────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────

BASE_URL    = "http://localhost:8000"
TENANT_ID   = "22000000-0000-0000-0000-000000000001"
TENANT_NAME = "CDL Network Test"
TENANT_SLUG = "cdl-network-test"

SCRIPT_DIR  = Path(__file__).parent.resolve()
ROOT_DIR    = SCRIPT_DIR.parent

def _find_input() -> Path:
    candidates = sorted(ROOT_DIR.glob("NetworkMembersA_updated_*.json"), reverse=True)
    if candidates:
        return candidates[0]
    candidates = sorted(ROOT_DIR.glob("NetworkMembersA*.json"), reverse=True)
    if candidates:
        return candidates[0]
    raise FileNotFoundError(
        f"No NetworkMembersA_updated_*.json found in {ROOT_DIR}. "
        "Place the file there and re-run."
    )

INPUT_JSON  = _find_input()
OUTPUT_FILE = ROOT_DIR / "cdl_matchmaking_results.xlsx"

def _read_pg_dsn() -> str:
    for candidate in [ROOT_DIR / ".env", SCRIPT_DIR / ".env", Path(".env")]:
        if candidate.exists():
            for line in candidate.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line.startswith("#"):
                    continue
                if line.lower().startswith("database_url="):
                    val = line.split("=", 1)[1].strip().strip('"').strip("'")
                    val = val.replace("postgresql+asyncpg://", "postgresql://")
                    val = val.replace("@localhost:", "@127.0.0.1:")
                    return val
    # Also try reading individual POSTGRES_* vars from .env
    pg_user = "delllo"
    pg_pass = "delllo_secret"
    pg_db   = "delllo_db"
    for candidate in [ROOT_DIR / ".env", SCRIPT_DIR / ".env", Path(".env")]:
        if candidate.exists():
            for line in candidate.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line.startswith("POSTGRES_USER="):
                    pg_user = line.split("=",1)[1].strip().strip('"').strip("'")
                elif line.startswith("POSTGRES_PASSWORD="):
                    pg_pass = line.split("=",1)[1].strip().strip('"').strip("'")
                elif line.startswith("POSTGRES_DB="):
                    pg_db   = line.split("=",1)[1].strip().strip('"').strip("'")
    return f"postgresql://{pg_user}:{pg_pass}@127.0.0.1:5432/{pg_db}"

PG_DSN       = _read_pg_dsn()
MAX_MATCHES  = 3
TX_TYPE      = "knowledge_transfer"
TIMEOUT_API  = 3000     # 50 minutes per request as requested

EXPECTED_MATCHES = [
    {"source": "Ajay Rathor",        "target": "Lee Williams",   "for": "Loan Trading Workflow Automation",                     "expected_score": 0.94},
    {"source": "Ajay Rathor",        "target": "Michael Parham", "for": "Structured Products Life Cycle Automation",            "expected_score": 0.91},
    {"source": "Ajay Rathor",        "target": "Hinesh Solanki", "for": "Front Office Technology",                              "expected_score": 0.89},
    {"source": "Ajay Rathor",        "target": "Steven Tian Yu", "for": "Front Office Technology",                              "expected_score": 0.90},
    {"source": "Priya Taneja-Rathor","target": "Katie Price",    "for": "Transaction Manager at GLAS",                         "expected_score": 0.88},
    {"source": "Priya Taneja-Rathor","target": "Brett Moody",    "for": "Corporate Trust / Structured Finance / Issuer Services","expected_score": 0.90},
]

# ─────────────────────────────────────────────────────────────────
#  ID helpers
# ─────────────────────────────────────────────────────────────────

_MONGO_NS = _uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

def name_to_uuid(name: str) -> str:
    return str(_uuid.uuid5(_MONGO_NS, f"{TENANT_ID}:{name}"))

# ─────────────────────────────────────────────────────────────────
#  Terminal helpers
# ─────────────────────────────────────────────────────────────────

results_log = []

def section(title):
    print(f"\n{'='*64}\n  {title}\n{'='*64}")

def info(msg):
    print(f"     {msg}")

def check(ok, msg):
    print(f"  {'OK' if ok else 'FAIL'}  {msg}")
    results_log.append((ok, msg))

# ─────────────────────────────────────────────────────────────────
#  Data helpers
# ─────────────────────────────────────────────────────────────────

def parse_json_col(val) -> list:
    if not val or (isinstance(val, float)):
        return []
    if isinstance(val, (list, dict)):
        return val if isinstance(val, list) else [val]
    try:
        return json.loads(val)
    except Exception:
        return []

def slug(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(text).lower()).strip('_')[:120]

def build_cv_text(row: dict) -> str:
    lines = []
    lines.append(f"{row['full_name']} -- {row.get('current_role','')} at {row.get('organisation_name','')}")
    lines.append("")
    lines.append(f"Location: {row.get('location','')}")
    lines.append(f"Seniority: {row.get('seniority','')}")
    lines.append(f"Role Family: {row.get('role_family','')}")
    lines.append(f"Persona: {row.get('persona_cluster','')}")
    lines.append(f"Primary Driver: {row.get('primary_driver_description','')}")
    lines.append(f"Secondary Driver: {row.get('secondary_driver_description','')}")
    lines.append("")
    for o in parse_json_col(row.get('offers_json')):
        if isinstance(o, dict): lines.append(f"OFFER: {o.get('summary','')}")
    for n in parse_json_col(row.get('needs_json')):
        if isinstance(n, dict): lines.append(f"NEED: {n.get('summary','')}")
    for s in parse_json_col(row.get('solutions_json')):
        if isinstance(s, dict): lines.append(f"SOLUTION: {s.get('summary','')}")
    for i in parse_json_col(row.get('insights_json')):
        if isinstance(i, dict): lines.append(f"INSIGHT: {i.get('summary','')}")
    return "\n".join(lines)

def build_facts(row: dict) -> List[tuple]:
    facts = []
    for o in parse_json_col(row.get('offers_json')):
        raw = o.get('summary','') if isinstance(o,dict) else str(o)
        if raw: facts.append(('offer', slug(raw), raw, 0.90))
    for n in parse_json_col(row.get('needs_json')):
        raw = n.get('summary','') if isinstance(n,dict) else str(n)
        if raw: facts.append(('need', slug(raw), raw, 0.90))
    for s in parse_json_col(row.get('solutions_json')):
        raw = s.get('summary','') if isinstance(s,dict) else str(s)
        if raw: facts.append(('skill', slug(raw), raw, 0.85))
    for i in parse_json_col(row.get('insights_json')):
        raw = i.get('summary','') if isinstance(i,dict) else str(i)
        if raw: facts.append(('topic', slug(raw), raw, 0.80))
    pd_desc = str(row.get('primary_driver_description',''))
    if pd_desc:
        facts.append(('domain', slug(row.get('primary_driver_id', pd_desc)), pd_desc, 0.95))
    sd_desc = str(row.get('secondary_driver_description',''))
    if sd_desc and sd_desc != pd_desc:
        facts.append(('domain', slug(row.get('secondary_driver_id', sd_desc)), sd_desc, 0.85))
    rf = str(row.get('role_family',''))
    if rf: facts.append(('skill', slug(rf), rf, 0.80))
    loc = str(row.get('location',''))
    if loc: facts.append(('location', slug(loc), loc, 0.95))
    persona = str(row.get('persona_cluster',''))
    if persona: facts.append(('topic', slug(persona), persona, 0.75))
    seen, deduped = set(), []
    for f in facts:
        k = (f[0], f[1])
        if k not in seen:
            seen.add(k)
            deduped.append(f)
    return deduped

# ─────────────────────────────────────────────────────────────────
#  Steps
# ─────────────────────────────────────────────────────────────────

def step_setup():
    section("Step 0 -- Setup CDL Tenant")
    with httpx.Client(timeout=30) as c:
        r = c.get(f"{BASE_URL}/health")
        check(r.status_code == 200, "API is healthy")
        if r.status_code != 200:
            print("  API not reachable. Is the docker stack running?")
            sys.exit(1)
        r = c.post(f"{BASE_URL}/v1/admin/wipe",
                   json={"tenant_id": TENANT_ID, "confirm": True})
        info("Previous data wiped" if r.status_code == 200 else "No prior data to wipe")
        r = c.post(f"{BASE_URL}/v1/tenants",
                   json={"tenant_id": TENANT_ID, "name": TENANT_NAME, "slug": TENANT_SLUG})
        if r.status_code in (200, 201):
            info(f"Tenant created: {TENANT_ID}")
        elif r.status_code == 409:
            info(f"Tenant already exists")
        else:
            info(f"Tenant status: {r.status_code}")


def step_create_users(records) -> Dict[str, str]:
    section(f"Step 1 -- Create {len(records)} Users")
    user_ids = {}
    with httpx.Client(timeout=30) as c:
        for i, row in enumerate(records):
            name  = row['full_name']
            uid   = name_to_uuid(name)
            email = re.sub(r'[^a-z0-9]', '.', name.lower()) + "@cdl-test.com"
            r = c.post(f"{BASE_URL}/v1/users", json={
                "user_id":      uid,
                "tenant_id":    TENANT_ID,
                "display_name": name,
                "email":        email,
                "headline":     f"{row.get('current_role','')} at {row.get('organisation_name','')}",
                "role":         "member",
                "status":       "active",
            })
            user_ids[name] = uid
            if (i + 1) % 20 == 0:
                check(r.status_code in (200,201), f"Users created: {i+1}/{len(records)}")
    info(f"Total users: {len(user_ids)}")
    return user_ids


def step_ingest(records, user_ids) -> Dict[str, str]:
    section("Step 2 -- Ingest CV Profiles")
    doc_ids = {}
    failed  = 0
    first_error = None
    with httpx.Client(timeout=60) as c:
        for i, row in enumerate(records):
            name = row['full_name']
            uid  = user_ids.get(name)
            if not uid:
                continue
            r = c.post(f"{BASE_URL}/v1/ingest/text", data={
                "tenant_id":   TENANT_ID,
                "user_id":     uid,
                "content":     build_cv_text(row),
                "source_type": "cv",
                "filename":    f"{name.replace(' ','_')}_cdl.txt",
                "embed":       "true",
            })
            if r.status_code == 200:
                doc_ids[name] = r.json().get("document_id","")
            else:
                failed += 1
                if first_error is None:
                    first_error = f"HTTP {r.status_code}: {r.text[:300]}"
            if (i + 1) % 20 == 0:
                print(f"  ->  Ingested {i+1}/{len(records)}...")
    if first_error:
        info(f"First ingestion error: {first_error}")
    check(failed == 0, f"All ingestions succeeded (failed: {failed})")
    info(f"Documents: {len(doc_ids)}")
    return doc_ids


def _try_pg_connect():
    """Try multiple connection strategies to reach Postgres."""
    dsns = [
        PG_DSN,
        "postgresql://delllo:delllo_secret@127.0.0.1:5432/delllo_db",
        "postgresql://delllo:delllo_secret@0.0.0.0:5432/delllo_db",
        "host=127.0.0.1 port=5432 dbname=delllo_db user=delllo password=delllo_secret",
    ]
    for dsn in dsns:
        try:
            conn = psycopg2.connect(dsn)
            info(f"Connected via: {re.sub(r':([^@/]+)@', ':****@', dsn)}")
            return conn
        except Exception:
            continue
    return None


VALID_FACT_TYPES = {'skill','domain','objective','offer','achievement','constraint','need','topic','location'}

def _seed_via_docker(records, user_ids, doc_ids) -> bool:
    import subprocess
    import tempfile
    import os

    info("Host connection blocked — using docker exec fallback...")

    sql_lines = ["BEGIN;"]
    count = 0
    skipped = 0
    for row in records:
        name   = row['full_name']
        uid    = user_ids.get(name)
        doc_id = doc_ids.get(name) or ""
        if not uid:
            continue
        for (ftype, canonical, raw, conf) in build_facts(row):
            if ftype not in VALID_FACT_TYPES:
                skipped += 1
                continue
            fact_id     = str(_uuid.uuid4())
            canonical_e = canonical.replace("'", "''")
            raw_e       = raw.replace("'", "''")
            doc_ref     = f"'{doc_id}'" if doc_id else "NULL"
            sql_lines.append(
                f"INSERT INTO extracted_facts "
                f"(fact_id,tenant_id,user_id,fact_type,canonical_value,raw_value,"
                f"confidence,visibility,source_document_id) VALUES "
                f"('{fact_id}','{TENANT_ID}','{uid}','{ftype}',"
                f"'{canonical_e}','{raw_e}',{conf},'match_engine_only',{doc_ref}) "
                f"ON CONFLICT (tenant_id,user_id,fact_type,canonical_value) "
                f"DO UPDATE SET raw_value=EXCLUDED.raw_value,"
                f"confidence=EXCLUDED.confidence;"
            )
            count += 1
    sql_lines.append("COMMIT;")

    if count == 0:
        info("No facts to seed")
        return True

    info(f"Writing {count} INSERT statements (skipped {skipped} invalid types)...")

    with tempfile.NamedTemporaryFile(mode='w', suffix='.sql',
                                     delete=False, encoding='utf-8') as f:
        f.write("\n".join(sql_lines))
        tmp_path = f.name

    try:
        cp = subprocess.run(
            ["docker", "cp", tmp_path, "delllo_postgres:/tmp/seed_facts.sql"],
            capture_output=True, text=True
        )
        if cp.returncode != 0:
            info(f"docker cp failed: {cp.stderr}")
            return False

        # Use ON_ERROR_STOP=0 so individual row errors don't abort the whole batch
        result = subprocess.run(
            ["docker", "exec", "delllo_postgres",
             "psql", "-U", "delllo", "-d", "delllo_db",
             "-f", "/tmp/seed_facts.sql"],
            capture_output=True, text=True
        )

        # Verify facts were actually written regardless of return code
        verify = subprocess.run(
            ["docker", "exec", "delllo_postgres",
             "psql", "-U", "delllo", "-d", "delllo_db", "-t", "-c",
             f"SELECT count(*) FROM extracted_facts WHERE tenant_id='{TENANT_ID}';"],
            capture_output=True, text=True
        )
        written = verify.stdout.strip()
        info(f"Verified facts in DB: {written}")

        if written and int(written) > 0:
            return True
        else:
            info(f"STDERR snippet: {result.stderr[:300]}")
            return False
    finally:
        os.unlink(tmp_path)


def step_seed_facts(records, user_ids, doc_ids):
    section("Step 3 -- Seed Facts -> Postgres + iKG")

    inserted = 0
    errors   = 0

    if HAS_PG:
        conn = _try_pg_connect()
    else:
        conn = None

    if conn is not None:
        # Direct host connection
        cur = conn.cursor()
        for i, row in enumerate(records):
            name   = row['full_name']
            uid    = user_ids.get(name)
            doc_id = doc_ids.get(name)
            if not uid:
                continue
            for (ftype, canonical, raw, conf) in build_facts(row):
                try:
                    cur.execute("""
                        INSERT INTO extracted_facts
                            (fact_id, tenant_id, user_id, fact_type,
                             canonical_value, raw_value, confidence,
                             visibility, source_document_id)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,'match_engine_only',%s)
                        ON CONFLICT (tenant_id, user_id, fact_type, canonical_value)
                            DO UPDATE SET raw_value=EXCLUDED.raw_value,
                                          confidence=EXCLUDED.confidence
                    """, (str(_uuid.uuid4()), TENANT_ID, uid, ftype,
                          canonical, raw, conf, doc_id))
                    inserted += 1
                except Exception as e:
                    errors += 1
                    conn.rollback()
            conn.commit()
            if (i + 1) % 20 == 0:
                print(f"  ->  Facts seeded: {i+1}/{len(records)} (total: {inserted})")
        cur.close()
        conn.close()
        check(errors == 0, f"Facts seeded via host (inserted: {inserted}, errors: {errors})")
    else:
        # Fallback: docker exec
        ok = _seed_via_docker(records, user_ids, doc_ids)
        check(ok, "Facts seeded via docker exec fallback")
        if not ok:
            info("Could not seed facts — matchmaking will run but scores may be low")

    section("Step 3b -- Sync iKG in Memgraph")
    synced = 0
    kfail  = 0
    with httpx.Client(timeout=30) as c:
        for name, uid in user_ids.items():
            r = c.post(f"{BASE_URL}/v1/ikg/upsert",
                       json={"user_id": uid, "tenant_id": TENANT_ID})
            if r.status_code == 200:
                synced += 1
            else:
                kfail += 1
            if synced % 20 == 0 and synced > 0:
                print(f"  ->  iKG synced: {synced}/{len(user_ids)}")
    check(kfail == 0, f"iKG sync (synced: {synced}, failed: {kfail})")


def step_signals(records, user_ids):
    section("Step 4 -- Post Intent Signals")
    posted = 0
    with httpx.Client(timeout=30) as c:
        for row in records:
            name = row['full_name']
            uid  = user_ids.get(name)
            if not uid:
                continue
            needs  = parse_json_col(row.get('needs_json'))
            intent = (needs[0].get('summary','') if needs and isinstance(needs[0],dict)
                      else str(row.get('primary_driver_description','')))
            if not intent:
                continue
            r = c.post(f"{BASE_URL}/v1/signals/intent", json={
                "tenant_id":   TENANT_ID,
                "user_id":     uid,
                "signal_type": "intent",
                "payload":     {"text": intent, "urgency": "medium"},
            })
            if r.status_code == 200:
                posted += 1
    check(posted > 0, f"Intent signals posted: {posted}/{len(records)}")


def step_matchmaking(records, user_ids) -> Dict[str, Any]:
    section(f"Step 5 -- Run Matchmaking (top {MAX_MATCHES} per user)")
    all_matches = {}
    failed      = 0
    active_users = list(user_ids.values())

    with httpx.Client(timeout=httpx.Timeout(TIMEOUT_API, connect=30)) as c:
        for i, row in enumerate(records):
            name = row['full_name']
            uid  = user_ids.get(name)
            if not uid:
                all_matches[name] = {"uid": None, "matches": [], "count": 0}
                continue

            r = c.post(f"{BASE_URL}/v1/matches/generate", json={
                "tenant_id":             TENANT_ID,
                "requesting_user_id":    uid,
                "transaction_types":     [TX_TYPE],
                "active_users":          active_users,
                "max_candidates":        MAX_MATCHES,
                "min_score":             0.0,
                "generate_explanations": True,
            })

            if r.status_code == 200:
                data    = r.json()
                matches = data.get("matches", [])
                all_matches[name] = {"uid": uid, "matches": matches, "count": len(matches)}
            else:
                failed += 1
                all_matches[name] = {"uid": uid, "matches": [], "count": 0}
                info(f"  WARN: {name}: HTTP {r.status_code}")

            if (i + 1) % 10 == 0:
                with_m = sum(1 for v in all_matches.values() if v["count"] > 0)
                print(f"  ->  Processed {i+1}/{len(records)}  (with matches: {with_m})")

    total = sum(v["count"] for v in all_matches.values())
    check(failed == 0, f"Matchmaking done (failed: {failed})")
    info(f"Total match pairs: {total}")
    return all_matches


# ─────────────────────────────────────────────────────────────────
#  Excel output
# ─────────────────────────────────────────────────────────────────

def step_save_excel(records, user_ids, all_matches):
    section("Step 6 -- Save Results to Excel")

    uid_to_name  = {v: k for k, v in user_ids.items()}
    name_to_row  = {r['full_name']: r for r in records}
    expected_set = {(e["source"], e["target"]) for e in EXPECTED_MATCHES}

    wb = Workbook()

    H  = PatternFill("solid", fgColor="1F4E79")
    AL = PatternFill("solid", fgColor="EBF3FA")
    HI = PatternFill("solid", fgColor="C6EFCE")
    ME = PatternFill("solid", fgColor="FFEB9C")
    LO = PatternFill("solid", fgColor="FFC7CE")
    EX = PatternFill("solid", fgColor="E2EFDA")
    CA = Alignment(horizontal="center", vertical="center")
    LA = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    BD = Border(left=Side(style="thin",color="D0D0D0"), right=Side(style="thin",color="D0D0D0"),
                top=Side(style="thin",color="D0D0D0"),   bottom=Side(style="thin",color="D0D0D0"))

    def hdr(ws, r, c, v):
        cell = ws.cell(r, c, value=v)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = H; cell.alignment = CA; cell.border = BD

    def std(ws, r, c, v, alt=False, left=False):
        cell = ws.cell(r, c, value=v)
        cell.font = Font(name="Arial", size=9)
        cell.fill = AL if alt else PatternFill()
        cell.alignment = LA if left else CA
        cell.border = BD
        return cell

    def sfill(v):
        if v is None: return PatternFill()
        if v >= 0.6:  return HI
        if v >= 0.3:  return ME
        return LO

    # Sheet 1: Match Results
    ws1 = wb.active
    ws1.title = "Match Results"
    h1 = ["Requester","Role","Organisation","Location",
          "Rank","Matched Person","Matched Role","Matched Org",
          "Score","Relevance","Complementarity","Timing","Proximity",
          "Evidence","Outcome Likelihood","Novelty","Expected?",
          "Why They Should Meet","Suggested Agenda","Opening Question",
          "Match ID"]
    for c,h in enumerate(h1,1): hdr(ws1,1,c,h)
    ws1.freeze_panes = "A2"

    rn = 2
    for idx, row in enumerate(records):
        name    = row['full_name']
        matches = all_matches.get(name,{}).get("matches",[])
        alt     = idx % 2 == 0
        if not matches:
            for c,v in enumerate([name,row.get('current_role',''),row.get('organisation_name',''),
                                   row.get('location',''),"--","No matches","","","","","","","","","","","","","","",""],1):
                std(ws1,rn,c,v,alt=alt,left=c<=4)
            rn += 1
            continue
        for rank, m in enumerate(matches, 1):
            mn   = uid_to_name.get(m.get("person_b",""), "")
            mr   = name_to_row.get(mn, {})
            bs   = m.get("score_breakdown") or {}
            sc   = m.get("score")
            iexp = (name, mn) in expected_set
            expl = m.get("explanation_text") or ""
            agnd = m.get("agenda_text") or ""
            oq   = m.get("opening_question") or ""
            vals = [name, row.get('current_role',''), row.get('organisation_name',''), row.get('location',''),
                    rank, mn, mr.get('current_role',''), mr.get('organisation_name',''),
                    sc,
                    bs.get("relevance"), bs.get("complementarity"), bs.get("timing"), bs.get("proximity"),
                    bs.get("evidence_strength"), bs.get("outcome_likelihood"), bs.get("novelty"),
                    "EXPECTED" if iexp else "",
                    expl, agnd, oq,
                    m.get("match_id","")]
            for c,v in enumerate(vals,1):
                cell = ws1.cell(rn, c)
                if 9 <= c <= 16:
                    cell.value = round(v,3) if isinstance(v,float) else v
                    cell.font = Font(name="Arial",size=9)
                    cell.fill = sfill(v) if isinstance(v,float) else PatternFill()
                    cell.alignment = CA; cell.border = BD
                    if isinstance(v,float): cell.number_format = "0.000"
                elif c in (18,19,20):
                    # explanation columns — left aligned, wrapped
                    std(ws1,rn,c,v,alt=alt,left=True)
                else:
                    std(ws1,rn,c,v,alt=alt,left=c in (1,2,3,6,7,8))
                if iexp and c==17:
                    ws1.cell(rn,c).fill = EX
                    ws1.cell(rn,c).font = Font(name="Arial",size=9,bold=True,color="375623")
            rn += 1

    for i,w in enumerate([24,30,22,14,6,24,30,22,8,8,8,8,8,8,8,8,12,50,40,40,36],1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    # Set row height to accommodate wrapped text in explanation columns
    for row_idx in range(2, rn):
        ws1.row_dimensions[row_idx].height = 60

    # Sheet 2: Expected Match Validation
    ws2 = wb.create_sheet("Expected Validation")
    h2  = ["Source Person","Target Person","Match Purpose","Expected Score",
           "RAIN Found?","RAIN Rank","RAIN Score","Score Diff","Result"]
    for c,h in enumerate(h2,1): hdr(ws2,1,c,h)

    for i, exp in enumerate(EXPECTED_MATCHES, 2):
        src, tgt    = exp["source"], exp["target"]
        fnd_rank    = None
        fnd_score   = None
        for rank, m in enumerate(all_matches.get(src,{}).get("matches",[]),1):
            if uid_to_name.get(m.get("person_b",""),"") == tgt:
                fnd_rank  = rank
                fnd_score = m.get("score")
                break
        found  = fnd_rank is not None
        diff   = round(fnd_score - exp["expected_score"],3) if fnd_score else None
        result = "MATCH FOUND" if found else "NOT FOUND"
        alt    = i % 2 == 0
        for c,v in enumerate([src,tgt,exp["for"],exp["expected_score"],
                               "Yes" if found else "No",
                               fnd_rank or "--",
                               round(fnd_score,3) if fnd_score else "--",
                               diff or "--", result],1):
            cell = std(ws2,i,c,v,alt=alt,left=c<=3)
            if c == 9:
                cell.fill = HI if found else LO
                cell.font = Font(name="Arial",size=9,bold=True)

    for i,w in enumerate([24,24,42,14,8,8,10,10,14],1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3: User Summary
    ws3 = wb.create_sheet("User Summary")
    h3  = ["Name","Role","Organisation","Location","Persona","Primary Driver",
           "Matches","Top Match","Top Score","User ID"]
    for c,h in enumerate(h3,1): hdr(ws3,1,c,h)

    for i, row in enumerate(records, 2):
        name    = row['full_name']
        matches = all_matches.get(name,{}).get("matches",[])
        tn      = uid_to_name.get(matches[0].get("person_b",""),"") if matches else ""
        ts      = matches[0].get("score") if matches else None
        alt     = i % 2 == 0
        for c,v in enumerate([name,row.get('current_role',''),row.get('organisation_name',''),
                               row.get('location',''),row.get('persona_cluster',''),
                               row.get('primary_driver_description',''),
                               len(matches),tn,
                               round(ts,3) if ts else None,
                               user_ids.get(name,"")],1):
            if c==9 and ts:
                cell = ws3.cell(i,c,value=v)
                cell.font=Font(name="Arial",size=9); cell.fill=sfill(v)
                cell.alignment=CA; cell.border=BD; cell.number_format="0.000"
            else:
                std(ws3,i,c,v,alt=alt,left=c in (1,2,3,6,8))

    for i,w in enumerate([24,32,22,14,24,50,10,24,10,36],1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4: Analytics
    ws4 = wb.create_sheet("Analytics")
    total   = len(records)
    matched = sum(1 for v in all_matches.values() if v["count"]>0)
    total_p = sum(v["count"] for v in all_matches.values())
    scores  = [m.get("score",0) for v in all_matches.values()
               for m in v.get("matches",[]) if m.get("score") is not None]
    avg_sc  = round(sum(scores)/len(scores),4) if scores else 0
    exp_fnd = sum(1 for e in EXPECTED_MATCHES
                  if any(uid_to_name.get(m.get("person_b",""))==e["target"]
                         for m in all_matches.get(e["source"],{}).get("matches",[])))

    rows4 = [
        ("CDL NETWORK TEST RESULTS",  None),
        ("Run Date",         datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Input File",       INPUT_JSON.name),
        ("Tenant ID",        TENANT_ID),
        ("",                 None),
        ("POPULATION",       None),
        ("Total Users",      total),
        ("Users With Matches", matched),
        ("Coverage %",       f"{matched/total*100:.1f}%"),
        ("",                 None),
        ("MATCH STATS",      None),
        ("Total Match Pairs",total_p),
        ("Avg Matches/User", f"{total_p/total:.1f}"),
        ("Average Score",    avg_sc),
        ("Max Score",        max(scores) if scores else 0),
        ("Min Score",        min(scores) if scores else 0),
        ("",                 None),
        ("EXPECTED VALIDATION", None),
        ("Expected Matches", len(EXPECTED_MATCHES)),
        ("Found by RAIN",    exp_fnd),
        ("Accuracy",         f"{exp_fnd/len(EXPECTED_MATCHES)*100:.0f}%"),
    ]
    for rn,(label,val) in enumerate(rows4,1):
        cl = ws4.cell(rn,1,value=label); cv = ws4.cell(rn,2,value=val)
        cl.border=BD; cv.border=BD
        if val is None:
            cl.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
            cl.fill=H; cl.alignment=LA
        else:
            cl.font=Font(name="Arial",bold=True,size=10); cl.alignment=LA
            cv.font=Font(name="Arial",size=10); cv.alignment=CA
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 42

    wb.save(OUTPUT_FILE)
    check(True, f"Saved -> {OUTPUT_FILE}")


# ─────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────

def main():
    print(f"\n  Delllo RAIN3.0 -- CDL Network Matchmaking Test")
    print(f"  API:     {BASE_URL}")
    print(f"  Tenant:  {TENANT_ID}")
    print(f"  Input:   {INPUT_JSON}")
    print(f"  Output:  {OUTPUT_FILE}")
    print(f"  Timeout: {TIMEOUT_API}s per request")
    print(f"  PG Mode: {'FAST (direct Postgres)' if HAS_PG else 'needs psycopg2-binary'}\n")

    data    = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    records = data["sheets"]["Sheet1"]["records"]

    # Deduplicate by full_name
    seen, unique = set(), []
    for r in records:
        if r["full_name"] not in seen:
            seen.add(r["full_name"])
            unique.append(r)
    records = unique
    info(f"Loaded {len(records)} unique users")
    check(len(records) > 0, "Dataset has users")

    t0 = time.time()
    step_setup()
    user_ids    = step_create_users(records)
    doc_ids     = step_ingest(records, user_ids)
    step_seed_facts(records, user_ids, doc_ids)
    step_signals(records, user_ids)
    all_matches = step_matchmaking(records, user_ids)
    step_save_excel(records, user_ids, all_matches)

    elapsed = time.time() - t0
    passed  = sum(1 for ok,_ in results_log if ok)
    failed  = sum(1 for ok,_ in results_log if not ok)
    total_m = sum(v["count"] for v in all_matches.values())

    section("Final Summary")
    info(f"Time:    {elapsed:.0f}s  ({elapsed/60:.1f} min)")
    info(f"Users:   {len(records)}")
    info(f"Matches: {total_m}")
    info(f"Checks:  {passed} passed / {failed} failed")
    print()
    if failed == 0:
        print("  All checks passed -- open cdl_matchmaking_results.xlsx")
    else:
        print(f"  {failed} check(s) failed -- see output above")
    print()


if __name__ == "__main__":
    main()